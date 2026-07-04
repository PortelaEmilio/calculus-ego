"""
Genera un XLSX de comparación IA vs anotación manual.

Formato de salida: misma ordenación de filas que el Excel manual, con columnas
IA y Manual intercaladas (primero IA, después Manual) y coloreado de diferencias.

Manejo de desajuste en número de personas:
  - Si IA detecta más personas → filas extra al final del bloque (fondo azul).
  - Si manual tiene más personas → filas extra al final del bloque (fondo amarillo).
"""

from __future__ import annotations

import io
import json
from collections import Counter
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string

# Conversión de píxeles ↔ unidades Excel (a 96 DPI, fuente por defecto Calibri 11)
_PX_TO_PT      = 0.75   # 1 px = 0.75 pt  (72 pt/in ÷ 96 px/in)
_PX_PER_CHAR   = 7      # px por carácter en Calibri 11 a 96 DPI

# Altura uniforme para todas las imágenes incrustadas (en px).
# El ancho varía manteniendo el aspect ratio original.
_IMG_TARGET_H  = 500
# Margen extra (px) para evitar redondeos que provoquen un desbordamiento de 1-2px
_IMG_PAD_PX    = 4

# ── Pares de columnas (IA, Manual, tipo) ────────────────────────────────────

COL_PAIRS: list[tuple[str, str, str]] = [
    ("IA Gen.",       "Man. Gen.",       "cat"),
    ("IA Edad",       "Man. Edad",       "cat"),
    ("IA Comport.",   "Man. Comport.",   "cat"),
    ("IA Activ.",     "Man. Activ.",     "cat"),
    ("IA Exp. Cp.",   "Man. Exp. Cp.",   "cat"),
    ("IA Ubic.",      "Man. Ubic.",      "cat"),
    ("IA Dist. Soc.", "Man. Dist. Soc.", "cat"),
    ("IA Maquill.",   "Man. Maquill.",   "bin"),
    ("IA Tattoos",    "Man. Tattoos",    "bin"),
    ("IA Bolsos",     "Man. Bolsos",     "bin"),
    ("IA Cints.",     "Man. Cints.",     "bin"),
    ("IA Joyas",      "Man. Joyas",      "bin"),
    ("IA Sombr.",     "Man. Sombr.",     "bin"),
    ("IA Gafas",      "Man. Gafas",      "bin"),
]

AI_COLS    = [p[0] for p in COL_PAIRS]
MAN_COLS   = [p[1] for p in COL_PAIRS]
BIN_PAIRS  = [(p[0], p[1]) for p in COL_PAIRS if p[2] == "bin"]

FIXED_LEFT_COLS  = ["Img ID", "Archivo", "Man. Nº pers.", "IA Nº pers.",
                     "Pers. Idx", "IA Pers. Idx"]
FIXED_RIGHT_COLS = ["IA Belleza", "Mismatches", "Notas",
                     "Ruta Img.", "Proc. At"]

PAIR_COLS = []
for _ai, _man, _ in COL_PAIRS:
    PAIR_COLS.extend([_ai, _man])

ALL_COLS          = FIXED_LEFT_COLS + PAIR_COLS + FIXED_RIGHT_COLS
ALL_COLS_WITH_IMG = ["Imagen"] + ALL_COLS

# ── Normalización ────────────────────────────────────────────────────────────

_BINARY_NORM = {
    "yes": "yes", "sí": "yes", "si": "yes",
    "no": "no",
    "1": "yes", "0": "no",
}


def _norm(val: object, binary: bool = False) -> str:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return "NA"
    s = str(val).strip().lower()
    if s in ("na", "nan", "none", ""):
        return "NA"
    if binary:
        return _BINARY_NORM.get(s, s)
    return s


def _mode_value(values: list) -> str:
    clean = [_norm(v) for v in values]
    clean = [v for v in clean if v != "NA"]
    if not clean:
        return "NA"
    return Counter(clean).most_common(1)[0][0]


# ── Extracción de datos IA desde JSON ────────────────────────────────────────

def _get_ordered_track_ids(result: dict) -> list:
    """Devuelve track_ids únicos ordenados por frame de primera aparición.

    Compatible con el formato antiguo (campo 'track_ids' en el resultado) y el
    formato nuevo (sin ese campo; track_ids extraídos de las listas de
    clasificación).
    """
    CAT_KEYS = [
        "gender_classifications", "age_classifications",
        "behaviour_classifications", "body_display_classifications",
        "activity_classifications",
    ]

    first_frame: dict = {}
    for cat in CAT_KEYS:
        for item in result.get(cat, []):
            tid = item.get("track_id")
            frame = item.get("frame", 0)
            if tid is not None and (tid not in first_frame or frame < first_frame[tid]):
                first_frame[tid] = frame

    # Formato antiguo: campo explícito
    explicit = result.get("track_ids")
    track_ids = list(explicit) if explicit else list(first_frame.keys())

    if not track_ids:
        return []

    def _key(tid):
        frame = first_frame.get(tid, 9999)
        if isinstance(tid, str) and tid.startswith("det_"):
            parts = tid.split("_")
            if len(parts) >= 3:
                try:
                    return (frame, int(parts[2]))
                except ValueError:
                    pass
        return (frame, 0 if isinstance(tid, int) else 9999)

    return sorted(track_ids, key=_key)


def _extract_accessories(item: dict) -> dict[str, bool]:
    """Normaliza el ítem de accesorios (formato nuevo con campos binarios o
    formato antiguo con lista 'accessories')."""
    if "makeup" in item or "tattoos" in item:
        return {
            "makeup":   bool(item.get("makeup",   0)),
            "tattoos":  bool(item.get("tattoos",  0)),
            "bags":     bool(item.get("bags",     0)),
            "belts":    bool(item.get("belts",    0)),
            "jewelry":  bool(item.get("jewelry",  0)),
            "headwear": bool(item.get("headwear", 0)),
            "eyewear":  bool(item.get("eyewear",  0)),
        }
    acc_list = [a.lower() for a in (item.get("accessories") or [])]
    return {
        "makeup":   any(k in acc_list for k in ("makeup", "maquillaje")),
        "tattoos":  any(k in acc_list for k in ("tattoos", "tatuajes")),
        "bags":     any(k in acc_list for k in ("bags", "bolsos")),
        "belts":    any(k in acc_list for k in ("belts", "cinturones")),
        "jewelry":  any(k in acc_list for k in ("jewelry", "jewellery", "joyas")),
        "headwear": any(k in acc_list for k in ("headwear", "hat", "sombrero")),
        "eyewear":  any(k in acc_list for k in ("eyewear", "glasses", "gafas")),
    }


def _extract_person_data(result: dict, track_id: object) -> dict:
    """Extrae las clasificaciones IA para un track_id concreto."""
    data: dict = {}

    def _pick(cat: str, field: str) -> str:
        items = [i for i in result.get(cat, []) if i.get("track_id") == track_id]
        return _mode_value([i.get(field) for i in items])

    data["IA Gen."]       = _pick("gender_classifications",        "gender")
    data["IA Edad"]       = _pick("age_classifications",           "age_group")
    data["IA Comport."]   = _pick("behaviour_classifications",     "behaviour")
    data["IA Activ."]     = _pick("activity_classifications",      "activity")
    data["IA Exp. Cp."]   = _pick("body_display_classifications",  "body_display")
    data["IA Ubic."]      = _pick("location_classifications",      "location")
    data["IA Dist. Soc."] = _pick("social_distance_classifications", "category")

    acc_items = [i for i in result.get("accessory_classifications", [])
                 if i.get("track_id") == track_id]
    if acc_items:
        agg = {k: False for k in ("makeup", "tattoos", "bags", "belts",
                                   "jewelry", "headwear", "eyewear")}
        for item in acc_items:
            for k, v in _extract_accessories(item).items():
                if v:
                    agg[k] = True
        data["IA Maquill."] = "yes" if agg["makeup"]   else "no"
        data["IA Tattoos"]  = "yes" if agg["tattoos"]  else "no"
        data["IA Bolsos"]   = "yes" if agg["bags"]     else "no"
        data["IA Cints."]   = "yes" if agg["belts"]    else "no"
        data["IA Joyas"]    = "yes" if agg["jewelry"]  else "no"
        data["IA Sombr."]   = "yes" if agg["headwear"] else "no"
        data["IA Gafas"]    = "yes" if agg["eyewear"]  else "no"
    else:
        for col in ["IA Maquill.", "IA Tattoos", "IA Bolsos", "IA Cints.",
                    "IA Joyas", "IA Sombr.", "IA Gafas"]:
            data[col] = "NA"

    beauty_items = [i for i in result.get("beauty_scores", [])
                    if i.get("track_id") == track_id]
    scores = []
    for i in beauty_items:
        raw = i.get("score")
        try:
            if raw is not None:
                scores.append(float(raw))
        except (ValueError, TypeError):
            pass
    data["IA Belleza"] = round(sum(scores) / len(scores), 1) if scores else "NA"

    return data


_EMPTY_AI: dict      = {col: "NA" for col in AI_COLS + ["IA Belleza"]}
_UNANALYZED_AI: dict = {col: "no analizado" for col in AI_COLS + ["IA Belleza"]}


def _find_json_fallback(img_id: str, json_path_str: object) -> Path | None:
    """Busca el JSON por img_id en los directorios de runs disponibles,
    ordenados de más reciente a más antiguo."""
    # Intentar directorio hermano del JSON original
    if json_path_str and not (isinstance(json_path_str, float) and pd.isna(json_path_str)):
        original = Path(str(json_path_str))
        # Escalar hasta encontrar un directorio 'runs'
        for parent in original.parents:
            if parent.name == "runs":
                runs_root = parent
                break
            candidate = parent.parent / "runs"
            if candidate.exists():
                runs_root = candidate
                break
        else:
            runs_root = None
    else:
        runs_root = None

    # Fallback: buscar 'runs' relativo al directorio de trabajo
    if runs_root is None or not runs_root.exists():
        cwd_runs = Path.cwd() / "validacion_imagenes" / "runs"
        if cwd_runs.exists():
            runs_root = cwd_runs

    if runs_root is None or not runs_root.exists():
        return None

    # Ordenar runs de más reciente a más antiguo y buscar el JSON
    fname = f"summary_{img_id}.json"
    run_dirs = sorted(runs_root.iterdir(), key=lambda p: p.name, reverse=True)
    for run_dir in run_dirs:
        if not run_dir.is_dir():
            continue
        for subdir in ["json", "."]:
            candidate = run_dir / subdir / fname
            if candidate.exists():
                return candidate
    return None


def _load_json_result(
    json_path_str: object,
    img_id: str = "",
) -> tuple[dict | None, int, list, int]:
    """Carga el JSON y devuelve (result, n_personas, track_ids_ordenados, n_personas_yolo).

    `n_personas` = personas procesadas por el VLM (= len(track_ids)).
    `n_personas_yolo` = detecciones totales de YOLO26, incluye las que se
    filtraron por BBOX_MIN_FRAME_RATIO. Cuando el JSON viene de un run previo
    al cambio (sin `persons_detected_yolo`), se devuelve igual a `n_personas`.

    Si la ruta original no existe, intenta localizar el JSON en los directorios
    de runs disponibles."""
    primary = None
    if json_path_str and not (isinstance(json_path_str, float) and pd.isna(json_path_str)):
        primary = Path(str(json_path_str))

    path = primary if (primary and primary.exists()) else _find_json_fallback(img_id, json_path_str)

    if path is None:
        return None, 0, [], 0
    try:
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
        result = data.get("result", data)
        track_ids = _get_ordered_track_ids(result)
        n_processed = len(track_ids)
        n_yolo = int(result.get("persons_detected_yolo", n_processed))
        return result, n_processed, track_ids, n_yolo
    except Exception as exc:
        print(f"  AVISO: error cargando {path.name}: {exc}")
        return None, 0, [], 0


# ── Construcción de filas de comparación ─────────────────────────────────────

def _count_mismatches(row: dict) -> int:
    n = 0
    for ai_col, man_col, kind in COL_PAIRS:
        ai_v  = _norm(row.get(ai_col,  ""), binary=(kind == "bin"))
        man_v = _norm(row.get(man_col, ""), binary=(kind == "bin"))
        if ai_v not in ("NA", "no analizado", "") and man_v not in ("NA", ""):
            if ai_v != man_v:
                n += 1
    return n


def build_comparison_rows(manual_df: pd.DataFrame) -> list[dict]:
    rows: list[dict] = []

    # Preservar orden de aparición de imágenes
    img_ids_ordered = list(dict.fromkeys(manual_df["Img ID"].astype(str).tolist()))

    for img_id in img_ids_ordered:
        img_mask   = manual_df["Img ID"].astype(str) == img_id
        img_rows   = manual_df[img_mask]
        n_manual   = len(img_rows)

        sample_row = img_rows.iloc[0]
        result, n_ai, track_ids, n_ai_yolo = _load_json_result(sample_row.get("Ruta JSON"), img_id)

        if result is None:
            print(f"  Sin JSON: {img_id}")

        n_total = max(n_manual, n_ai)

        for i in range(n_total):
            man_row  = img_rows.iloc[i] if i < n_manual else None
            track_id = track_ids[i] if i < n_ai else None

            if i < n_manual and i < n_ai:
                row_type = "matched"
            elif i < n_manual:
                row_type = "manual_only"
            else:
                row_type = "ai_only"

            meta = man_row if man_row is not None else sample_row

            row: dict = {
                "Img ID":        img_id,
                "Archivo":       str(meta.get("Archivo", "")),
                "Man. Nº pers.": str(sample_row.get("Man. Nº pers.", "")),
                "IA Nº pers.":   n_ai_yolo,
                "Pers. Idx":     str(man_row.get("Pers. Idx", "")) if man_row is not None else "",
                "IA Pers. Idx":  i if i < n_ai else "",
                "Notas":         str(man_row.get("Notas", "")) if man_row is not None else "",
                "Ruta Img.":     str(meta.get("Ruta Img.", "")),
                "Proc. At":      str(meta.get("Proc. At", "")),
                "_row_type":     row_type,
            }

            # Datos IA
            if track_id is not None and result is not None:
                row.update(_extract_person_data(result, track_id))
            elif result is not None:
                # Imagen analizada pero sin detecciones → NA
                row.update(_EMPTY_AI)
            else:
                # Sin JSON: no se ha analizado
                row.update(_UNANALYZED_AI)

            # Datos manuales
            if man_row is not None:
                for col in MAN_COLS:
                    raw = man_row.get(col)
                    row[col] = "" if (raw is None or (isinstance(raw, float) and pd.isna(raw))) else str(raw)
            else:
                for col in MAN_COLS:
                    row[col] = ""

            row["_is_first_row"] = (i == 0)
            row["Mismatches"] = _count_mismatches(row) if row_type == "matched" else ""
            rows.append(row)

    return rows


# ── Incrustación de imágenes ──────────────────────────────────────────────────

def _resolve_runs_root(ruta_json_str: object) -> Path | None:
    """Resuelve la raíz `runs/` a partir de la ruta del JSON o del cwd."""
    if ruta_json_str and not (isinstance(ruta_json_str, float) and pd.isna(ruta_json_str)):
        for parent in Path(str(ruta_json_str)).parents:
            if parent.name == "runs":
                return parent
    cwd_runs = Path.cwd() / "validacion_imagenes" / "runs"
    return cwd_runs if cwd_runs.exists() else None


def _find_annotated_image(img_id: str, ruta_json_str: object) -> Path | None:
    """Busca la imagen anotada `<imgID>_annotated.jpg` en los directorios de runs.
    Devuelve la ruta del run más reciente que contenga la imagen, o None."""
    runs_root = _resolve_runs_root(ruta_json_str)
    if runs_root is None or not runs_root.exists():
        return None

    fname = f"{img_id}_annotated.jpg"
    for run_dir in sorted(runs_root.iterdir(), key=lambda p: p.name, reverse=True):
        if not run_dir.is_dir():
            continue
        for subdir in ["annotated", "."]:
            c = run_dir / subdir / fname
            if c.exists():
                return c
    return None


def _find_person_annotated_crop(
    img_id: str,
    person_idx: int | str,
    ruta_json_str: object,
) -> Path | None:
    """Busca el recorte anotado por persona generado por
    `save_person_annotated_crop` en `<run>/annotated/person_crops_annotated/`.

    El nombre de archivo sigue el patrón
    `<imgID>_person_<idx>_track_<trackID>_annotated.jpg`; el track_id se
    desconoce a la hora de exportar, así que se hace glob por idx."""
    if person_idx == "" or person_idx is None:
        return None
    try:
        idx_int = int(person_idx)
    except (TypeError, ValueError):
        return None

    runs_root = _resolve_runs_root(ruta_json_str)
    if runs_root is None or not runs_root.exists():
        return None

    pattern = f"{img_id}_person_{idx_int}_track_*_annotated.jpg"
    fallback_pattern = f"{img_id}_person_{idx_int}_annotated.jpg"

    for run_dir in sorted(runs_root.iterdir(), key=lambda p: p.name, reverse=True):
        if not run_dir.is_dir():
            continue
        for subdir in ["annotated/person_crops_annotated", "person_crops_annotated"]:
            base = run_dir / subdir
            if not base.exists():
                continue
            matches = sorted(base.glob(pattern))
            if matches:
                return matches[0]
            fb = base / fallback_pattern
            if fb.exists():
                return fb
    return None


def _load_full_image(img_path: Path, target_h: int = _IMG_TARGET_H) -> tuple[bytes, int, int] | None:
    """Carga la imagen anotada y la redimensiona a una altura uniforme target_h px,
    preservando el aspect ratio original (la anchura varía).
    Retorna (bytes_jpeg, w_px, h_px) o None si falla."""
    try:
        from PIL import Image as PILImage
        with PILImage.open(img_path) as pil_img:
            if pil_img.mode not in ("RGB", "L"):
                pil_img = pil_img.convert("RGB")
            w, h = pil_img.size
            scale = target_h / h
            new_w = max(1, int(w * scale))
            new_h = target_h
            pil_img = pil_img.resize((new_w, new_h), PILImage.LANCZOS)
            buf = io.BytesIO()
            pil_img.save(buf, format="JPEG", quality=90, optimize=True, dpi=(96, 96))
            return buf.getvalue(), new_w, new_h
    except Exception as exc:
        print(f"  AVISO imagen {img_path.name}: {exc}")
        return None


def _embed_image(ws, thumb_bytes: bytes, w_px: int, h_px: int,
                 row_idx: int, col_idx: int) -> None:
    """Inserta la imagen con tamaño fijo (w_px × h_px) en la celda (col_idx, row_idx).
    OneCellAnchor + ext en EMU: el tamaño es exacto, sin deformar."""
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    from openpyxl.utils.units import pixels_to_EMU

    buf = io.BytesIO(thumb_bytes)
    xl_img = XLImage(buf)
    xl_img.width  = w_px
    xl_img.height = h_px

    anchor = OneCellAnchor()
    anchor._from = AnchorMarker(col=col_idx - 1, colOff=0,
                                row=row_idx - 1, rowOff=0)
    anchor.ext = XDRPositiveSize2D(cx=pixels_to_EMU(w_px),
                                   cy=pixels_to_EMU(h_px))
    xl_img.anchor = anchor
    ws.add_image(xl_img)


# ── Escritura del XLSX ────────────────────────────────────────────────────────

def write_comparison_xlsx(
    rows: list[dict],
    output_path: Path,
    include_images: bool = True,
) -> None:
    FILL_MATCH    = PatternFill("solid", fgColor="C8E6C9")
    FILL_MISMATCH = PatternFill("solid", fgColor="FFCDD2")
    FILL_NA_MAN   = PatternFill("solid", fgColor="EEEEEE")
    FILL_NO_AI    = PatternFill("solid", fgColor="F5F5F5")
    FILL_AI_ONLY  = PatternFill("solid", fgColor="E3F2FD")
    FILL_MAN_ONLY = PatternFill("solid", fgColor="FFF9C4")
    FILL_HDR_META = PatternFill("solid", fgColor="37474F")
    FILL_HDR_AI   = PatternFill("solid", fgColor="1565C0")
    FILL_HDR_MAN  = PatternFill("solid", fgColor="E65100")
    FILL_HDR_IMG  = PatternFill("solid", fgColor="4A148C")
    FONT_WHITE    = Font(color="FFFFFF", bold=True)
    ALIGN_CENTER  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ALIGN_LEFT    = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    col_list = ALL_COLS_WITH_IMG if include_images else ALL_COLS
    col_map  = {col: idx for idx, col in enumerate(col_list, 1)}

    wb = Workbook()
    ws = wb.active
    ws.title = "Comparación IA vs Manual"
    ws.row_dimensions[1].height = 36

    # Columna imagen — el ancho se ajustará al final, tras conocer el máximo.
    img_col_letter: str | None = None
    if include_images and "Imagen" in col_map:
        img_col_letter = get_column_letter(col_map["Imagen"])
    max_img_w_px = 0  # se actualizará al insertar cada imagen

    # ── Cabecera ──────────────────────────────────────────────────────────────
    for col_name, col_idx in col_map.items():
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.alignment = ALIGN_CENTER
        if col_name == "Imagen":
            cell.fill = FILL_HDR_IMG
            cell.font = FONT_WHITE
        elif col_name in AI_COLS or col_name in ("IA Belleza",
                                                   "IA Nº pers.", "IA Pers. Idx"):
            cell.fill = FILL_HDR_AI
            cell.font = FONT_WHITE
        elif col_name in MAN_COLS or col_name in ("Man. Nº pers.", "Pers. Idx"):
            cell.fill = FILL_HDR_MAN
            cell.font = FONT_WHITE
        else:
            cell.fill = FILL_HDR_META
            cell.font = FONT_WHITE

    # ── Datos ─────────────────────────────────────────────────────────────────
    for row_offset, row_data in enumerate(rows, 2):
        row_type = row_data.get("_row_type",    "matched")
        is_first = row_data.get("_is_first_row", False)

        for col_name, col_idx in col_map.items():
            if col_name == "Imagen":
                continue
            val  = row_data.get(col_name, "")
            cell = ws.cell(row=row_offset, column=col_idx,
                           value=val if val != "" else None)
            cell.alignment = ALIGN_LEFT

        # Imagen incrustada: preferir el recorte anotado de esta persona;
        # si no existe (p.ej. corrida antigua sin per-person), usar la imagen
        # anotada completa solo en la primera fila del grupo.
        if include_images and img_col_letter:
            img_id_str = str(row_data.get("Img ID", ""))
            ruta_json  = row_data.get("Ruta JSON", row_data.get("Ruta Img.", ""))
            ia_idx     = row_data.get("IA Pers. Idx", "")

            person_crop_path = _find_person_annotated_crop(img_id_str, ia_idx, ruta_json) \
                if ia_idx != "" and ia_idx is not None else None

            chosen: Path | None = person_crop_path
            if chosen is None and is_first:
                annotated = _find_annotated_image(img_id_str, ruta_json)
                chosen = annotated if annotated else (
                    Path(row_data.get("Ruta Img.", "")) if row_data.get("Ruta Img.") else None
                )

            if chosen and chosen.exists():
                result = _load_full_image(chosen)
                if result:
                    img_bytes, w_px, h_px = result
                    _embed_image(ws, img_bytes, w_px, h_px, row_offset,
                                 column_index_from_string(img_col_letter))
                    # Altura con un pequeño margen para evitar desbordamiento por redondeo
                    ws.row_dimensions[row_offset].height = (h_px + _IMG_PAD_PX) * _PX_TO_PT
                    if w_px > max_img_w_px:
                        max_img_w_px = w_px

        # Fondo de fila para personas extra
        if row_type == "ai_only":
            for col_idx in col_map.values():
                ws.cell(row=row_offset, column=col_idx).fill = FILL_AI_ONLY
        elif row_type == "manual_only":
            for col_idx in col_map.values():
                ws.cell(row=row_offset, column=col_idx).fill = FILL_MAN_ONLY

        # Color por par de columnas (filas emparejadas)
        if row_type == "matched":
            for ai_col, man_col, kind in COL_PAIRS:
                ai_v  = _norm(row_data.get(ai_col,  ""), binary=(kind == "bin"))
                man_v = _norm(row_data.get(man_col, ""), binary=(kind == "bin"))

                if man_v in ("NA", ""):
                    fill = FILL_NA_MAN
                elif ai_v in ("NA", "no analizado", ""):
                    fill = FILL_NO_AI
                elif ai_v == man_v:
                    fill = FILL_MATCH
                else:
                    fill = FILL_MISMATCH

                ws.cell(row=row_offset, column=col_map[ai_col]).fill  = fill
                ws.cell(row=row_offset, column=col_map[man_col]).fill = fill

    # ── Anchos de columna ─────────────────────────────────────────────────────
    MIN_W, MAX_W = 8, 30
    for col_name, col_idx in col_map.items():
        if col_name == "Imagen":
            continue
        letter  = get_column_letter(col_idx)
        max_len = max(
            (len(str(ws.cell(row=r, column=col_idx).value or ""))
             for r in range(1, ws.max_row + 1)),
            default=MIN_W,
        )
        ws.column_dimensions[letter].width = min(max_len + 2, MAX_W)

    # Ajuste final del ancho de la columna "Imagen" al máximo encontrado.
    # Fórmula inversa de Excel: col_w = (px - 5) / 7 (Calibri 11 @ 96 DPI).
    # Se añaden _IMG_PAD_PX de margen para evitar desbordamientos por redondeo.
    if include_images and img_col_letter and max_img_w_px > 0:
        target_px = max_img_w_px + _IMG_PAD_PX
        col_w = max(0.0, (target_px - 5) / _PX_PER_CHAR)
        ws.column_dimensions[img_col_letter].width = min(col_w, 255)

    freeze = "B2" if include_images else "A2"
    ws.freeze_panes = freeze
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


# ── Entrada pública ───────────────────────────────────────────────────────────

def generar_xlsx_comparacion(
    manual_xlsx: str | Path,
    output_xlsx: str | Path | None = None,
    include_images: bool = True,
) -> Path:
    manual_xlsx = Path(manual_xlsx)
    if output_xlsx is None:
        output_xlsx = manual_xlsx.parent / "comparacion_ia_manual.xlsx"
    output_xlsx = Path(output_xlsx)

    print(f"Cargando: {manual_xlsx.name}")
    df = pd.read_excel(manual_xlsx, engine="openpyxl", sheet_name=0)
    df["Img ID"] = df["Img ID"].astype(str)
    print(f"  {len(df)} filas, {df['Img ID'].nunique()} imágenes únicas")

    print("Construyendo filas de comparación…")
    rows = build_comparison_rows(df)

    matched  = sum(1 for r in rows if r["_row_type"] == "matched")
    ai_only  = sum(1 for r in rows if r["_row_type"] == "ai_only")
    man_only = sum(1 for r in rows if r["_row_type"] == "manual_only")
    print(f"  Total filas: {len(rows)}  "
          f"(emparejadas: {matched}, solo-IA: {ai_only}, solo-manual: {man_only})")

    print(f"Escribiendo Excel en: {output_xlsx}")
    write_comparison_xlsx(rows, output_xlsx, include_images=include_images)
    print("Listo.")
    return output_xlsx
